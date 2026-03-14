import csv
import io
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz

# =========================================================
# CONFIGURAÇÃO
# =========================================================
st.set_page_config(page_title="ANVISA - Registro por Lotes", layout="wide")

DATA_DIR = Path("data")
MED_CSV_DEFAULT = DATA_DIR / "TA_PRECO_MEDICAMENTO.csv"
PROD_CSV_DEFAULTS = [
    DATA_DIR / "TA_PRODUTO_SAUDE_SITE.csv",
    DATA_DIR / "TA_CONSULTA_PRODUTOS_SAUDE.CSV",
]

LOTES_PADRAO = [f"L0{i}" for i in range(1, 9)]

FILL_HEADER = PatternFill(fill_type="solid", start_color="C8E6C9", end_color="C8E6C9")
FILL_ALTERADO = PatternFill(fill_type="solid", start_color="FFF59D", end_color="FFF59D")
FILL_OK = PatternFill(fill_type="solid", start_color="C8E6C9", end_color="C8E6C9")

STOPWORDS = {
    "DE", "DA", "DO", "DAS", "DOS", "COM", "SEM", "PARA", "PCT", "UND", "UN",
    "CX", "KIT", "LT", "L", "EM", "E", "OU", "AO", "A", "O", "X", "C", "P",
    "COMERCIAL", "PRODUTO", "NOME", "SOLUCAO", "SUSPENSAO", "ORAL", "INJETAVEL"
}


# =========================================================
# FUNÇÕES BÁSICAS
# =========================================================
def normalizar_texto(txt):
    if txt is None:
        return ""
    txt = str(txt).strip().upper()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = txt.replace("\n", " ").replace("\r", " ")
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def apenas_digitos(txt):
    if txt is None:
        return ""
    return re.sub(r"\D", "", str(txt))


def limpar_desc(desc):
    desc = normalizar_texto(desc)

    trocas = {
        "SOL. ORAL.": "SOLUCAO ORAL",
        "SOL ORAL": "SOLUCAO ORAL",
        "SUSP. ORAL.": "SUSPENSAO ORAL",
        "SUSP ORAL": "SUSPENSAO ORAL",
        "COMPRIMIDO": "CPR",
        "COMPRIMIDOS": "CPR",
        "CAPSULA": "CAP",
        "CAPSULAS": "CAP",
        "AMPOLA": "AMP",
        "AMPOLAS": "AMP",
        "FRASCO": "FR",
        "FRASCOS": "FR",
        "BISNAGA": "BISN",
        "MILILITROS": "ML",
        "MILILITRO": "ML",
        "GRAMAS": "G",
        "GRAMA": "G",
        "MICROGRAMAS": "MCG",
        "MICROGRAMA": "MCG",
        "UNIDADES INTERNACIONAIS": "UI",
        "INJETÁVEL": "INJETAVEL",
        "SUSPENSÃO": "SUSPENSAO",
        "SOLUÇÃO": "SOLUCAO",
        "P/": "PARA ",
        " C/ ": " COM ",
    }

    for de, para in trocas.items():
        desc = desc.replace(de, para)

    desc = re.sub(r"[.,;:()+\-/%]", " ", desc)
    desc = re.sub(r"\s+", " ", desc).strip()
    return desc


def tokenizar(txt):
    txt = limpar_desc(txt)
    tokens = []
    for token in txt.split():
        if len(token) < 2:
            continue
        if token in STOPWORDS:
            continue
        tokens.append(token)
    return list(dict.fromkeys(tokens))


def eh_linha_vazia_ou_total(valor_desc):
    v = normalizar_texto(valor_desc)
    if not v:
        return True

    termos_ruins = {
        "TOTAL",
        "TOTAL DO LOTE",
        "SUBTOTAL",
        "VALOR TOTAL",
        "TOTAL GERAL",
    }
    return v in termos_ruins


def parece_registravel(desc):
    d = limpar_desc(desc)

    palavras_chave = [
        "MG", "ML", "MCG", "G", "UI",
        "SOLUCAO", "SUSPENSAO", "INJETAVEL",
        "CPR", "CAP", "AMP", "FR", "BOLSA", "BISN",
        "CREME", "POMADA", "XAROPE", "GEL", "SOLUCAO ORAL",
        "SERINGA", "CATETER", "EQUIPO", "LUVA", "CURATIVO",
        "KIT", "TESTE", "REAGENTE", "AGULHA", "ABAIXADOR",
    ]

    return any(p in d for p in palavras_chave)


def primeira_coluna_existente(df, candidatos):
    norm = {normalizar_texto(c): c for c in df.columns}

    for cand in candidatos:
        cand_norm = normalizar_texto(cand)
        for real_norm, real_col in norm.items():
            if cand_norm == real_norm:
                return real_col

    for cand in candidatos:
        cand_norm = normalizar_texto(cand)
        for real_norm, real_col in norm.items():
            if cand_norm in real_norm:
                return real_col

    return None


# =========================================================
# LEITOR ROBUSTO DE CSV
# =========================================================
def detectar_separador(linhas):
    candidatos = [";", "\t", ",", "|"]
    contagens = {sep: sum(l.count(sep) for l in linhas[:50]) for sep in candidatos}
    return max(contagens, key=contagens.get)


def detectar_linha_cabecalho_csv(linhas, sep):
    melhor_idx = 0
    melhor_score = -1

    for idx, linha in enumerate(linhas[:150]):
        partes = [p.strip().strip('"') for p in linha.split(sep)]
        joined = " ".join(partes).upper()
        score = 0

        if "REGISTRO" in joined:
            score += 3
        if "PRODUTO" in joined or "NOME_COMERCIAL" in joined or "NOME COMERCIAL" in joined:
            score += 2
        if "APRESENT" in joined or "NOME_TECNICO" in joined:
            score += 1
        if "LABORAT" in joined or "DETENTOR" in joined or "EMPRESA" in joined:
            score += 2
        if len(partes) >= 5:
            score += 1

        if score > melhor_score:
            melhor_score = score
            melhor_idx = idx

    return melhor_idx


def ler_csv_robusto(origem):
    if hasattr(origem, "getvalue"):
        bruto = origem.getvalue()
        texto = bruto.decode("utf-8", errors="ignore")
    else:
        with open(origem, "rb") as f:
            texto = f.read().decode("utf-8", errors="ignore")

    linhas = texto.splitlines()
    if not linhas:
        return pd.DataFrame()

    sep = detectar_separador(linhas)
    hdr_idx = detectar_linha_cabecalho_csv(linhas, sep)

    reader = csv.reader(linhas[hdr_idx:], delimiter=sep, quotechar='"')
    rows = list(reader)
    if not rows:
        return pd.DataFrame()

    header = rows[0]
    data = [r for r in rows[1:] if any(str(c).strip() for c in r)]

    n = len(header)
    data_fix = []
    for r in data:
        if len(r) < n:
            r = r + [""] * (n - len(r))
        elif len(r) > n:
            r = r[:n]
        data_fix.append(r)

    return pd.DataFrame(data_fix, columns=header).fillna("")


# =========================================================
# EXCEL - CABEÇALHO E COLUNAS
# =========================================================
def achar_linha_cabecalho_excel(ws):
    melhores = []

    for r in range(1, min(ws.max_row, 20) + 1):
        valores = [normalizar_texto(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 15) + 1)]
        score = 0

        for v in valores:
            if v == "ITEM":
                score += 2
            if "DESCRICAO" in v:
                score += 3
            if v == "MARCA":
                score += 2
            if v in ("UN", "UND", "UNIDADE"):
                score += 2
            if "QUANT" in v:
                score += 1
            if "ANVISA" in v:
                score += 2

        melhores.append((score, r))

    melhores.sort(reverse=True)
    if not melhores:
        return None

    return melhores[0][1] if melhores[0][0] >= 5 else None


def mapear_colunas_excel(ws, header_row):
    mapa = {}

    for c in range(1, ws.max_column + 1):
        nome = normalizar_texto(ws.cell(header_row, c).value)
        nome = nome.replace("Nº", "N ").replace("NO ", "N ").replace("N.", "N ")

        if nome == "ITEM":
            mapa["ITEM"] = c
        elif "DESCRICAO" in nome:
            mapa["DESCRICAO"] = c
        elif nome == "MARCA":
            mapa["MARCA"] = c
        elif "REGISTRO" in nome and "ANVISA" in nome:
            mapa["ANVISA"] = c
        elif nome in ("UN", "UND", "UNIDADE"):
            mapa["UN"] = c
        elif "QUANT" in nome:
            mapa["QUANT"] = c

    return mapa


def inserir_coluna_anvisa(ws, header_row):
    mapa = mapear_colunas_excel(ws, header_row)

    if "ANVISA" in mapa:
        return mapa["ANVISA"], False

    if "MARCA" not in mapa or "UN" not in mapa:
        return None, False

    col_marca = mapa["MARCA"]
    ws.insert_cols(col_marca + 1)
    ws.cell(header_row, col_marca + 1).value = "Nº REGISTRO NA ANVISA"
    ws.cell(header_row, col_marca + 1).fill = FILL_HEADER
    return col_marca + 1, True


# =========================================================
# PREPARAÇÃO DAS BASES
# =========================================================
def preparar_base(df, tipo):
    if tipo == "med":
        col_prod = primeira_coluna_existente(df, ["PRODUTO", "NOME PRODUTO", "NOME_DO_PRODUTO", "NOME COMERCIAL"])
        col_apres = primeira_coluna_existente(df, ["APRESENTAÇÃO", "APRESENTACAO", "DESCRICAO APRESENTACAO"])
        col_reg = primeira_coluna_existente(df, ["REGISTRO", "NUMERO REGISTRO", "NUMERO DO REGISTRO"])
        col_marca = primeira_coluna_existente(df, ["LABORATÓRIO", "LABORATORIO", "DETENTOR", "EMPRESA", "RAZAO SOCIAL"])
        col_sit = primeira_coluna_existente(df, ["TIPO DE PRODUTO (STATUS DO PRODUTO)", "STATUS", "SITUACAO"])

        base = pd.DataFrame({
            "produto": df[col_prod] if col_prod else "",
            "apresentacao": df[col_apres] if col_apres else "",
            "registro": df[col_reg] if col_reg else "",
            "marca": df[col_marca] if col_marca else "",
            "situacao": df[col_sit] if col_sit else "",
        })
        base["produto_apresentacao"] = (
            base["produto"].astype(str) + " " + base["apresentacao"].astype(str)
        ).map(limpar_desc)

    else:
        col_prod = primeira_coluna_existente(df, [
            "NOME COMERCIAL", "NOME_COMERCIAL", "NOME TECNICO", "NOME_TECNICO", "PRODUTO"
        ])
        col_reg = primeira_coluna_existente(df, [
            "NUMERO_REGISTRO_CADASTRO", "NUMERO REGISTRO", "REGISTRO", "NUMERO CADASTRO"
        ])
        col_marca = primeira_coluna_existente(df, [
            "DETENTOR_REGISTRO_CADASTRO", "DETENTOR DO REGISTRO", "DETENTOR", "EMPRESA", "RAZAO SOCIAL"
        ])
        col_sit = primeira_coluna_existente(df, [
            "VALIDADE_REGISTRO_CADASTRO", "SITUACAO", "STATUS"
        ])

        base = pd.DataFrame({
            "produto": df[col_prod] if col_prod else "",
            "registro": df[col_reg] if col_reg else "",
            "marca": df[col_marca] if col_marca else "",
            "situacao": df[col_sit] if col_sit else "",
        })
        base["produto_apresentacao"] = base["produto"].map(limpar_desc)

    base["registro_dig"] = base["registro"].map(apenas_digitos)
    base["marca_norm"] = base["marca"].map(normalizar_texto)
    base["situacao_norm"] = base["situacao"].map(normalizar_texto)
    base = base[base["produto_apresentacao"].astype(str).str.len() > 2].copy().reset_index(drop=True)
    base["tokens"] = base["produto_apresentacao"].map(tokenizar)
    return base


def build_index(df):
    indice = defaultdict(list)
    for i, tokens in enumerate(df["tokens"]):
        for tok in set(tokens[:12]):
            indice[tok].append(i)
    return indice


@st.cache_resource(show_spinner=False)
def carregar_bases_e_indices():
    if not MED_CSV_DEFAULT.exists():
        raise FileNotFoundError("Arquivo não encontrado: data/TA_PRECO_MEDICAMENTO.csv")

    prod_path = None
    for p in PROD_CSV_DEFAULTS:
        if p.exists():
            prod_path = p
            break

    if prod_path is None:
        raise FileNotFoundError(
            "Arquivo não encontrado em data/: use TA_PRODUTO_SAUDE_SITE.csv ou TA_CONSULTA_PRODUTOS_SAUDE.CSV"
        )

    med_raw = ler_csv_robusto(MED_CSV_DEFAULT)
    prod_raw = ler_csv_robusto(prod_path)

    med = preparar_base(med_raw, "med")
    prod = preparar_base(prod_raw, "prod")

    med_idx = build_index(med)
    prod_idx = build_index(prod)

    return med, prod, med_idx, prod_idx, str(prod_path.name)


# =========================================================
# MATCH
# =========================================================
def filtrar_ativos(base_df):
    if "situacao_norm" not in base_df.columns:
        return base_df

    ativos = base_df[
        base_df["situacao_norm"].astype(str).str.contains(
            "NOVO|SIMILAR|GENERICO|ESPECIFICO|BIOLOGICO|FITOTERAPICO|DINAMIZADO|VIGENTE|ATIV|VALID|REGULAR",
            na=False,
            regex=True,
        )
    ]
    return ativos if not ativos.empty else base_df


def score_linha(desc_item, marca_item, produto_base, marca_base):
    s_desc = fuzz.token_set_ratio(desc_item, produto_base)
    s_marca = fuzz.ratio(marca_item, marca_base) if marca_item and marca_base else 0
    return (s_desc * 0.85) + (s_marca * 0.15)


def candidate_indices(desc_item, indice, limit=250):
    tokens = tokenizar(desc_item)
    if not tokens:
        return []

    counts = Counter()
    for tok in tokens[:8]:
        for idx in indice.get(tok, []):
            counts[idx] += 1

    if not counts:
        return []

    return [idx for idx, _ in counts.most_common(limit)]


def melhor_match(desc_item, marca_item, base_df, indice, score_minimo=70):
    desc_limpa = limpar_desc(desc_item)
    marca_norm = normalizar_texto(marca_item)

    if not desc_limpa:
        return None

    candidatos_idx = candidate_indices(desc_limpa, indice, 250)
    if not candidatos_idx:
        return None

    melhor = None
    melhor_score = -1

    for idx in candidatos_idx:
        row = base_df.iloc[idx]
        sc = score_linha(
            desc_item=desc_limpa,
            marca_item=marca_norm,
            produto_base=row.get("produto_apresentacao", ""),
            marca_base=row.get("marca_norm", ""),
        )
        if sc > melhor_score:
            melhor_score = sc
            melhor = row

    if melhor is None:
        return None

    score_final = round(melhor_score, 2)
    if score_final < score_minimo:
        return None

    return {
        "registro": melhor.get("registro_dig", ""),
        "marca_base": melhor.get("marca", ""),
        "situacao": melhor.get("situacao", ""),
        "score": score_final,
    }


def buscar_registro(desc_item, marca_item, med_df, prod_df, med_idx, prod_idx):
    if parece_registravel(desc_item):
        med = melhor_match(desc_item, marca_item, med_df, med_idx, score_minimo=68)
        if med and med["registro"]:
            return {"fonte": "MEDICAMENTO", **med}

    prod = melhor_match(desc_item, marca_item, prod_df, prod_idx, score_minimo=72)
    if prod and prod["registro"]:
        return {"fonte": "PRODUTO_SAUDE", **prod}

    med2 = melhor_match(desc_item, marca_item, med_df, med_idx, score_minimo=72)
    if med2 and med2["registro"]:
        return {"fonte": "MEDICAMENTO_FALLBACK", **med2}

    return None


# =========================================================
# PROCESSAMENTO EXCEL
# =========================================================
def contar_linhas_processaveis(wb, lotes_alvo):
    total = 0
    for aba in wb.sheetnames:
        if aba not in lotes_alvo:
            continue
        ws = wb[aba]
        header_row = achar_linha_cabecalho_excel(ws)
        if not header_row:
            continue
        mapa = mapear_colunas_excel(ws, header_row)
        col_desc = mapa.get("DESCRICAO")
        if not col_desc:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            desc = ws.cell(r, col_desc).value
            if not eh_linha_vazia_ou_total(desc):
                total += 1
    return total


def processar_excel(workbook_bytes, lotes_alvo, med_df, prod_df, med_idx, prod_idx, trocar_marca=True, modo_conferencia=False):
    wb = load_workbook(io.BytesIO(workbook_bytes))
    auditoria = []

    total_linhas = contar_linhas_processaveis(wb, lotes_alvo)
    progresso = st.progress(0, text="Iniciando processamento...")
    contador = 0

    for aba in wb.sheetnames:
        if aba not in lotes_alvo:
            continue

        ws = wb[aba]
        header_row = achar_linha_cabecalho_excel(ws)

        if not header_row:
            auditoria.append({
                "aba": aba,
                "linha": "",
                "item": "",
                "descricao": "",
                "marca_original": "",
                "marca_final": "",
                "registro_encontrado": "",
                "registro_planilha": "",
                "status": "Cabeçalho não identificado",
                "score": "",
                "fonte": "",
            })
            continue

        col_anvisa, _ = inserir_coluna_anvisa(ws, header_row)
        mapa = mapear_colunas_excel(ws, header_row)

        if col_anvisa is None:
            auditoria.append({
                "aba": aba,
                "linha": "",
                "item": "",
                "descricao": "",
                "marca_original": "",
                "marca_final": "",
                "registro_encontrado": "",
                "registro_planilha": "",
                "status": "Não foi possível inserir/achar a coluna ANVISA",
                "score": "",
                "fonte": "",
            })
            continue

        mapa = mapear_colunas_excel(ws, header_row)
        col_item = mapa.get("ITEM")
        col_desc = mapa.get("DESCRICAO")
        col_marca = mapa.get("MARCA")
        col_un = mapa.get("UN")
        col_anvisa = mapa.get("ANVISA")

        if not col_desc or not col_marca or not col_un or not col_anvisa:
            auditoria.append({
                "aba": aba,
                "linha": "",
                "item": "",
                "descricao": "",
                "marca_original": "",
                "marca_final": "",
                "registro_encontrado": "",
                "registro_planilha": "",
                "status": "Colunas essenciais ausentes",
                "score": "",
                "fonte": "",
            })
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            item = ws.cell(r, col_item).value if col_item else None
            desc = ws.cell(r, col_desc).value
            marca_original = ws.cell(r, col_marca).value
            reg_planilha = ws.cell(r, col_anvisa).value

            if eh_linha_vazia_ou_total(desc):
                continue

            contador += 1
            if total_linhas > 0:
                progresso.progress(
                    min(contador / total_linhas, 1.0),
                    text=f"Processando {aba} - linha {r} ({contador}/{total_linhas})"
                )

            resultado = buscar_registro(desc, marca_original, med_df, prod_df, med_idx, prod_idx)

            if resultado:
                registro_encontrado = resultado["registro"]
                marca_sugerida = resultado["marca_base"]
                score = resultado["score"]
                fonte = resultado["fonte"]

                if modo_conferencia:
                    status = "CONFERE" if apenas_digitos(reg_planilha) == registro_encontrado else "DIVERGENTE"
                else:
                    ws.cell(r, col_anvisa).value = registro_encontrado
                    ws.cell(r, col_anvisa).fill = FILL_OK
                    status = "PREENCHIDO"

                    if trocar_marca:
                        marca_atual_norm = normalizar_texto(marca_original)
                        marca_nova_norm = normalizar_texto(marca_sugerida)
                        if marca_nova_norm and marca_nova_norm != marca_atual_norm:
                            ws.cell(r, col_marca).value = marca_sugerida
                            ws.cell(r, col_marca).fill = FILL_ALTERADO
                            status = "PREENCHIDO + MARCA AJUSTADA"

                auditoria.append({
                    "aba": aba,
                    "linha": r,
                    "item": item,
                    "descricao": desc,
                    "marca_original": marca_original,
                    "marca_final": marca_sugerida if marca_sugerida else marca_original,
                    "registro_encontrado": registro_encontrado,
                    "registro_planilha": apenas_digitos(reg_planilha),
                    "status": status,
                    "score": score,
                    "fonte": fonte,
                })
            else:
                auditoria.append({
                    "aba": aba,
                    "linha": r,
                    "item": item,
                    "descricao": desc,
                    "marca_original": marca_original,
                    "marca_final": marca_original,
                    "registro_encontrado": "",
                    "registro_planilha": apenas_digitos(reg_planilha),
                    "status": "NAO ENCONTRADO",
                    "score": "",
                    "fonte": "",
                })

    progresso.progress(1.0, text="Finalizando e gerando arquivos...")

    if "AUDITORIA_ANVISA" in wb.sheetnames:
        del wb["AUDITORIA_ANVISA"]

    ws_aud = wb.create_sheet("AUDITORIA_ANVISA")
    headers = [
        "aba", "linha", "item", "descricao",
        "marca_original", "marca_final",
        "registro_encontrado", "registro_planilha",
        "status", "score", "fonte"
    ]

    for c, h in enumerate(headers, 1):
        ws_aud.cell(1, c).value = h
        ws_aud.cell(1, c).fill = FILL_HEADER

    for i, row in enumerate(auditoria, start=2):
        for c, h in enumerate(headers, 1):
            ws_aud.cell(i, c).value = row.get(h, "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, pd.DataFrame(auditoria)


# =========================================================
# CONFERÊNCIA PDF
# =========================================================
def extrair_linhas_pdf(pdf_bytes):
    linhas = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pagina_idx, pagina in enumerate(pdf.pages, start=1):
            texto = pagina.extract_text() or ""
            for linha in texto.splitlines():
                linha_limpa = linha.strip()
                if linha_limpa:
                    linhas.append({"pagina": pagina_idx, "linha_pdf": linha_limpa})

    return pd.DataFrame(linhas)


def conferir_pdf(pdf_bytes, med_df, prod_df, med_idx, prod_idx):
    linhas_df = extrair_linhas_pdf(pdf_bytes)
    resultados = []

    if linhas_df.empty:
        return pd.DataFrame([{
            "pagina": "",
            "linha_pdf": "",
            "descricao_identificada": "",
            "registro_encontrado": "",
            "status": "PDF sem texto extraível"
        }])

    padrao_reg = re.compile(r"\b\d{9,13}\b")
    progresso = st.progress(0, text="Lendo PDF...")

    for i, (_, row) in enumerate(linhas_df.iterrows(), start=1):
        linha = row["linha_pdf"]
        pagina = row["pagina"]

        progresso.progress(min(i / len(linhas_df), 1.0), text=f"Conferindo PDF... linha {i}/{len(linhas_df)}")

        if len(linha) < 8:
            continue

        registro_existente = ""
        m = padrao_reg.search(linha)
        if m:
            registro_existente = m.group(0)

        if not parece_registravel(linha):
            continue

        encontrado = buscar_registro(linha, "", med_df, prod_df, med_idx, prod_idx)
        if encontrado:
            reg = encontrado["registro"]
            status = "LOCALIZADO"
            if registro_existente:
                status = "CONFERE" if registro_existente == reg else "DIVERGENTE"

            resultados.append({
                "pagina": pagina,
                "linha_pdf": linha,
                "descricao_identificada": linha,
                "registro_encontrado": reg,
                "registro_existente_pdf": registro_existente,
                "status": status,
                "score": encontrado["score"],
                "fonte": encontrado["fonte"],
            })

    if not resultados:
        return pd.DataFrame([{
            "pagina": "",
            "linha_pdf": "",
            "descricao_identificada": "",
            "registro_encontrado": "",
            "registro_existente_pdf": "",
            "status": "Nenhuma linha aproveitável encontrada no PDF"
        }])

    return pd.DataFrame(resultados)


# =========================================================
# INTERFACE
# =========================================================
st.title("ANVISA - preenchimento e conferência por lotes")
st.write(
    "Processa planilhas com múltiplas abas, trabalha apenas com os lotes escolhidos "
    "e insere/preenche a coluna **Nº REGISTRO NA ANVISA** quando ela não existir."
)

with st.sidebar:
    st.subheader("Lotes")
    lotes_escolhidos = st.multiselect("Lotes a processar", options=LOTES_PADRAO, default=LOTES_PADRAO)
    trocar_marca = st.checkbox("Trocar MARCA se a marca ativa encontrada for diferente", value=True)

    st.markdown("---")
    st.subheader("Bases ANVISA")
    st.caption("Esta versão usa por padrão os CSVs locais já salvos na pasta data/.")

modo = st.radio("Escolha o modo", ["Preencher Excel", "Conferir Excel", "Conferir PDF"], horizontal=True)

if modo in ("Preencher Excel", "Conferir Excel"):
    arquivo_principal = st.file_uploader("Envie a planilha Excel (.xlsx)", type=["xlsx"])
else:
    arquivo_principal = st.file_uploader("Envie o PDF para conferência (.pdf)", type=["pdf"])

if st.button("Executar", type="primary"):
    try:
        med_df, prod_df, med_idx, prod_idx, nome_base_prod = carregar_bases_e_indices()

        st.info(
            f"Bases carregadas com sucesso. Medicamentos: {len(med_df):,} registros | "
            f"Produto saúde: {len(prod_df):,} registros | Base produto saúde usada: {nome_base_prod}"
        )

        if arquivo_principal is None:
            st.error("Envie o arquivo principal antes de executar.")
            st.stop()

        if modo == "Preencher Excel":
            saida, aud = processar_excel(
                workbook_bytes=arquivo_principal.getvalue(),
                lotes_alvo=lotes_escolhidos,
                med_df=med_df,
                prod_df=prod_df,
                med_idx=med_idx,
                prod_idx=prod_idx,
                trocar_marca=trocar_marca,
                modo_conferencia=False,
            )
            st.success("Processamento concluído.")
            st.dataframe(aud, use_container_width=True)

            st.download_button(
                "Baixar planilha processada",
                data=saida.getvalue(),
                file_name="planilha_processada_anvisa.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.download_button(
                "Baixar auditoria CSV",
                data=aud.to_csv(index=False).encode("utf-8-sig"),
                file_name="auditoria_anvisa.csv",
                mime="text/csv",
            )

        elif modo == "Conferir Excel":
            _, aud = processar_excel(
                workbook_bytes=arquivo_principal.getvalue(),
                lotes_alvo=lotes_escolhidos,
                med_df=med_df,
                prod_df=prod_df,
                med_idx=med_idx,
                prod_idx=prod_idx,
                trocar_marca=False,
                modo_conferencia=True,
            )
            st.success("Conferência da planilha concluída.")
            st.dataframe(aud, use_container_width=True)
            st.download_button(
                "Baixar relatório de conferência",
                data=aud.to_csv(index=False).encode("utf-8-sig"),
                file_name="conferencia_excel_anvisa.csv",
                mime="text/csv",
            )

        elif modo == "Conferir PDF":
            conf_pdf = conferir_pdf(
                pdf_bytes=arquivo_principal.getvalue(),
                med_df=med_df,
                prod_df=prod_df,
                med_idx=med_idx,
                prod_idx=prod_idx,
            )
            st.success("Conferência do PDF concluída.")
            st.dataframe(conf_pdf, use_container_width=True)
            st.download_button(
                "Baixar relatório do PDF",
                data=conf_pdf.to_csv(index=False).encode("utf-8-sig"),
                file_name="conferencia_pdf_anvisa.csv",
                mime="text/csv",
            )

    except Exception as e:
        st.error(f"Erro: {e}")
